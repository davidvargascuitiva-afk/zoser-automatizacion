import zipfile
import re
import os
import time
import openpyxl
import pandas as pd
from copy import copy as _copy_obj
from datetime import timedelta
from typing import List
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from app.models.sensor import Sensor
from app.models.proyecto import ProyectoConfig

# Mapping: template drawing file → target name in output
_USERSHAPES_MAP = {
    'xl/drawings/drawing2.xml': 'xl/drawings/us_gt.xml',
    'xl/drawings/drawing4.xml': 'xl/drawings/us_fallastem.xml',
    'xl/drawings/drawing6.xml': 'xl/drawings/us_ghr.xml',
    'xl/drawings/drawing8.xml': 'xl/drawings/us_fallashr.xml',
}
_CHART_RELS_TEMPLATE = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
    '<Relationship Id="rId3" '
    'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chartUserShapes" '
    'Target="{target}"/>'
    '</Relationships>'
)

# Chart line colors cycling through 5 lightness variants × 6 accents = 30 styles
# (covers up to 25 sensors with unique styles, 26-30 cycle back gracefully)
_CHART_LINE_STYLES = []
for _lm in [None, 60000, 40000, 20000, 10000]:
    for _acc in range(1, 7):
        _inner = (
            f'<a:lumMod val="{_lm}"/>' if _lm else ''
        )
        _CHART_LINE_STYLES.append(
            f'<a:ln w="28575" cap="rnd" cmpd="sng" algn="ctr">'
            f'<a:solidFill><a:schemeClr val="accent{_acc}">'
            f'{_inner}<a:shade val="95000"/><a:satMod val="105000"/>'
            f'</a:schemeClr></a:solidFill>'
            f'<a:prstDash val="solid"/><a:round/></a:ln><a:effectLst/>'
        )

# ─────────────────────────────────────────────────────────────────
#  STYLE HELPERS
# ─────────────────────────────────────────────────────────────────

def _copy_cell_style(src, dst) -> None:
    """Copy all formatting from src cell to dst without touching the value."""
    try:
        if src.has_style:
            dst.font         = _copy_obj(src.font)
            dst.border       = _copy_obj(src.border)
            dst.fill         = _copy_obj(src.fill)
            dst.number_format = src.number_format
            dst.alignment    = _copy_obj(src.alignment)
    except Exception:
        pass


def _capturar_estilo_fila(ws, row: int, cols: range) -> dict:
    """Snapshot cell styles for a row (before insert_rows shifts things)."""
    return {
        col: {
            'font':          _copy_obj(ws.cell(row=row, column=col).font)          if ws.cell(row=row, column=col).has_style else None,
            'border':        _copy_obj(ws.cell(row=row, column=col).border)        if ws.cell(row=row, column=col).has_style else None,
            'fill':          _copy_obj(ws.cell(row=row, column=col).fill)          if ws.cell(row=row, column=col).has_style else None,
            'number_format': ws.cell(row=row, column=col).number_format            if ws.cell(row=row, column=col).has_style else 'General',
            'alignment':     _copy_obj(ws.cell(row=row, column=col).alignment)     if ws.cell(row=row, column=col).has_style else None,
        }
        for col in cols
    }


def _aplicar_estilo_fila(ws, row: int, snapshot: dict) -> None:
    """Apply captured style snapshot to a row."""
    for col, s in snapshot.items():
        dst = ws.cell(row=row, column=col)
        try:
            if s['font']:          dst.font          = s['font']
            if s['border']:        dst.border        = s['border']
            if s['fill']:          dst.fill          = s['fill']
            if s['alignment']:     dst.alignment     = s['alignment']
            dst.number_format = s['number_format']
        except Exception:
            pass


def _estilizar_col_sensor(ws, pos: int, ref_pos: int = 9) -> None:
    """Copy column style (width + all relevant row ranges) from ref_pos to pos in a T/HR sheet."""
    ref_col = _t_data_col(ref_pos)
    dst_col = _t_data_col(pos)
    ref_ltr = get_column_letter(ref_col)
    dst_ltr = get_column_letter(dst_col)

    # Column width
    if ref_ltr in ws.column_dimensions:
        ws.column_dimensions[dst_ltr].width = ws.column_dimensions[ref_ltr].width

    # Metadata rows 10-16
    for r in range(10, 17):
        _copy_cell_style(ws.cell(row=r, column=ref_col), ws.cell(row=r, column=dst_col))

    # Data rows 17-305 — use row 17 as template (all data rows share the same style)
    ref_data = ws.cell(row=T_HR_DATA_ROW_START, column=ref_col)
    for r in range(T_HR_DATA_ROW_START, T_HR_DATA_ROW_START + MAX_PRIMARIOS):
        _copy_cell_style(ref_data, ws.cell(row=r, column=dst_col))

    # Stat rows 306-309
    for r in _STAT_ROWS_PER_SENSOR:
        _copy_cell_style(ws.cell(row=r, column=ref_col), ws.cell(row=r, column=dst_col))


COL_DATETIME        = 2
PRIMARIOS_ROW_START = 7
FALLAS_ROW_START    = 7
MAX_PRIMARIOS       = 289
MAX_FALLAS          = 60
SHEET_PRIMARIOS     = 'Primarios'
T_HR_DATA_ROW_START = 17
T_HR_COL_DATETIME   = 3
_STAT_ROWS_PER_SENSOR = [306, 307, 308, 309]   # MAX, MIN, AVG, DIST por sensor
_STAT_ROWS_GLOBAL     = [311, 312, 313]         # AVG global, MIN global, MAX global


def _cl(col_idx: int) -> str:
    return get_column_letter(col_idx)


def _prim_temp_col(sensor_pos: int) -> int:
    """1-based column in Primarios for sensor_pos temperature (starts at C=3)."""
    return 2 + sensor_pos


def _prim_hum_col(sensor_pos: int, n: int) -> int:
    """1-based column in Primarios for sensor_pos humidity (after temp block + skip)."""
    return n + 3 + sensor_pos


def _t_data_col(sensor_pos: int) -> int:
    """1-based column in T/HR sheet for sensor_pos data (starts at D=4)."""
    return 3 + sensor_pos


def _resolver_sheet(wb, candidatos: list):
    nombres = wb.sheetnames
    for c in candidatos:
        if c in nombres:
            return c
    return None


def llenar_plantilla(
    ruta_plantilla: str,
    ruta_salida: str,
    sensores: List[Sensor],
    timestamps: pd.Series,
    df_fallas_temp: pd.DataFrame,
    df_fallas_hum: pd.DataFrame,
    config: ProyectoConfig,
) -> None:
    n = config.num_sensores
    wb = openpyxl.load_workbook(ruta_plantilla, keep_vba=False)
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    # Expand Análisis if more than 9 sensors — copy row-31 style to new rows
    if n > 9:
        nombre_a = _resolver_sheet(wb, ['Análisis', 'Analisis', 'Análisis'])
        ws_a = wb[nombre_a] if nombre_a else None
        if ws_a:
            extra = n - 9
            snap = _capturar_estilo_fila(ws_a, 31, range(1, 16))
            # openpyxl does NOT shift merged ranges on insert_rows — only cell content
            # shifts. Every merged range with min_row >= 32 must be unmerged before the
            # insert and re-merged at (original + extra) rows afterwards, or the primary
            # cell stays at the old row (now empty) and the secondary cell (which has the
            # shifted content) can't display anything.
            merges_to_shift = []
            for mr in list(ws_a.merged_cells.ranges):
                if mr.min_row >= 32:
                    merges_to_shift.append(
                        (mr.min_row, mr.max_row, mr.min_col, mr.max_col))
                    ws_a.unmerge_cells(str(mr))
            ws_a.insert_rows(32, extra)
            for new_row in range(32, 32 + extra):
                _aplicar_estilo_fila(ws_a, new_row, snap)
                ws_a.row_dimensions[new_row].height = 13.5
            # Ensure column widths match for all sensor positions beyond original template range
            for pos in range(10, n + 1):
                ws_a.column_dimensions[get_column_letter(3 + pos)].width = 13.0
            # Re-apply every collected merge at its shifted position
            for min_row, max_row, min_col, max_col in merges_to_shift:
                ws_a.merge_cells(
                    start_row=min_row + extra, start_column=min_col,
                    end_row=max_row + extra, end_column=max_col)

    _llenar_primarios(wb, sensores, timestamps, n)
    _llenar_hoja_T(wb, sensores, timestamps, n)
    _llenar_hoja_HR(wb, sensores, timestamps, n)

    sheet_ft = _resolver_sheet(wb, ['Fallas Tem', 'Fallas Tem '])
    sheet_fh = _resolver_sheet(wb, ['Fallas HR', 'Fallas HR '])
    if sheet_ft:
        _llenar_fallas(wb, sheet_ft, df_fallas_temp, es_temp=True)
        _escribir_nombres_fallas(wb, sheet_ft, config.tipo_prueba)
    if sheet_fh:
        _llenar_fallas(wb, sheet_fh, df_fallas_hum, es_temp=False)
        _escribir_nombres_fallas(wb, sheet_fh, config.tipo_prueba)

    _llenar_analisis(wb, config, n, sensores)

    wb.save(ruta_salida)
    wb.close()

    _inyectar_cuadros_info(ruta_plantilla, ruta_salida, config)
    _patch_chart_n_sensores(ruta_salida, n)


# ─────────────────────────────────────────────────────────────────
#  PRIMARIOS
# ─────────────────────────────────────────────────────────────────

def _llenar_primarios(wb, sensores: List[Sensor], timestamps: pd.Series, n: int) -> None:
    ws = wb[SHEET_PRIMARIOS]
    rows = min(len(timestamps), MAX_PRIMARIOS)

    # Capture reference header style (blue fill) from template before any changes
    _ref_fill    = _copy_obj(ws.cell(row=5, column=3).fill)
    _ref_font    = _copy_obj(ws.cell(row=5, column=3).font)
    _header_font = openpyxl.styles.Font(name='Arial', size=8, bold=True)
    _data_font   = openpyxl.styles.Font(name='Arial', size=8, bold=False)
    _no_fill     = PatternFill()

    # Clear rows 1-6 across the entire possible range (max of template n=9 and current n)
    # This removes stale labels for unused columns when n < 9, and preps new cols when n > 9
    _max_clear_col = max(_prim_hum_col(9, 9), _prim_hum_col(n, n)) + 3
    for col in range(3, _max_clear_col):
        for r in range(1, 7):
            try:
                cell = ws.cell(row=r, column=col)
                cell.value = None
                if r >= 5:
                    cell.fill = _no_fill
            except Exception:
                pass

    # Write metadata (rows 1-4, Arial 8 bold) and headers (rows 5-6) for active sensors
    for sensor in sensores:
        pos = sensor.posicion
        if pos < 1 or pos > n:
            continue
        tc = _prim_temp_col(pos)
        hc = _prim_hum_col(pos, n)
        for col in (tc, hc):
            for r, val in [(1, sensor.nombre),
                           (2, 'Registrador Temperatura y Humedad'),
                           (3, sensor.serial),
                           (4, 'MultiChannel')]:
                cell = ws.cell(row=r, column=col)
                cell.value = val
                cell.font  = _copy_obj(_header_font)
        for r, val in [(5, f'Posición {pos}'), (6, 'Temperature (°C)')]:
            cell = ws.cell(row=r, column=tc)
            cell.value = val
            cell.fill  = _copy_obj(_ref_fill)
            cell.font  = _copy_obj(_ref_font)
        for r, val in [(5, f'Posición {pos}'), (6, 'Humedad (% RH)')]:
            cell = ws.cell(row=r, column=hc)
            cell.value = val
            cell.fill  = _copy_obj(_ref_fill)
            cell.font  = _copy_obj(_ref_font)

    _DT_FMT = 'DD/MM/YYYY h:mm am/pm'

    # Set column widths for all active sensor columns
    _skip_col = n + 3  # skip col between temp and hum blocks (L for n=9)
    for pos in range(1, n + 1):
        ws.column_dimensions[_cl(_prim_temp_col(pos))].width = 13.6640625 if pos == 1 else 13.0
        ws.column_dimensions[_cl(_prim_hum_col(pos, n))].width = 13.6640625 if pos == 1 else 13.0
    ws.column_dimensions[_cl(_skip_col)].width = 8.109375

    # Fix row-6 borders and hide unused columns for N < 9.
    # The N=9 template has a medium bottom border on row 6 across ALL 21 columns (A-U).
    # For N<9, columns beyond the last active col remain empty but keep that border,
    # creating a visible black line. Fix: remove bottom border from unused cols,
    # and ensure every active col has the correct border (the old N=9 skip col L
    # loses its border when the clear loop wipes its fill, so we restore it too).
    # Additionally hide the unused data columns so the Primarios sheet doesn't show
    # empty ghost columns with template cell formatting in rows 7-295.
    if n < 9:
        from openpyxl.styles import Border, Side
        _med  = Side(style='medium')
        _none = Side(style=None)
        last_active = max(_prim_temp_col(n), _prim_hum_col(n, n))
        # 1. Restore border on every active sensor column (value or skip)
        active_cols = (
            [_skip_col] +
            [_prim_temp_col(p) for p in range(1, n + 1)] +
            [_prim_hum_col(p, n) for p in range(1, n + 1)]
        )
        for col in active_cols:
            cell = ws.cell(row=6, column=col)
            try:
                b = cell.border
                cell.border = Border(
                    left=b.left, right=b.right, top=b.top,
                    bottom=_med
                )
            except Exception:
                pass
        # 2. Remove bottom border AND hide unused cols (last_active+1 … template max col 21)
        for col in range(last_active + 1, 22):
            cell = ws.cell(row=6, column=col)
            try:
                b = cell.border
                cell.border = Border(
                    left=b.left, right=b.right, top=b.top,
                    bottom=_none
                )
            except Exception:
                pass
            ltr = _cl(col)
            ws.column_dimensions[ltr].hidden = True
            ws.column_dimensions[ltr].width  = 0

    # Pre-apply number_format + font to ALL data cells (rows 7-295) for all N positions.
    # Done as a column-pass before the data loop so every cell gets consistent styling
    # regardless of whether the sensor has data (blanks should still be formatted).
    for pos in range(1, n + 1):
        tc = _prim_temp_col(pos)
        hc = _prim_hum_col(pos, n)
        for r in range(PRIMARIOS_ROW_START, PRIMARIOS_ROW_START + rows):
            t_cell = ws.cell(row=r, column=tc)
            h_cell = ws.cell(row=r, column=hc)
            t_cell.number_format = '0.00'
            t_cell.font = _copy_obj(_data_font)
            h_cell.number_format = '0.0'
            h_cell.font = _copy_obj(_data_font)

    for i in range(rows):
        row = PRIMARIOS_ROW_START + i
        cell = ws.cell(row=row, column=COL_DATETIME)
        cell.value = timestamps.iloc[i].to_pydatetime()
        cell.number_format = _DT_FMT
        for sensor in sensores:
            pos = sensor.posicion
            if pos < 1 or pos > n or sensor.datos is None or i >= len(sensor.datos):
                continue
            ws.cell(row=row, column=_prim_temp_col(pos)).value = _safe_temp(sensor.datos['temperatura'].iloc[i])
            ws.cell(row=row, column=_prim_hum_col(pos, n)).value = _safe_hum(sensor.datos['humedad'].iloc[i])


# ─────────────────────────────────────────────────────────────────
#  T SHEET (temperature + GT chart data)
# ─────────────────────────────────────────────────────────────────

def _llenar_hoja_T(wb, sensores: List[Sensor], timestamps: pd.Series, n: int) -> None:
    ws = wb['T']
    rows = min(len(timestamps), MAX_PRIMARIOS)

    # ── Estilos: copiar columna sensor-9 a columnas nuevas (n > 9) ─────────────
    for pos in range(10, n + 1):
        _estilizar_col_sensor(ws, pos, ref_pos=9)

    # ── Limpiar y ocultar columnas no usadas cuando n < 9 ──────────────────────
    # Clear value + fill + border for EVERY row in the column (1-330).
    # Listing specific rows (1-16, 306-309, 311-313, 321) misses gaps like row 310
    # and rows 314-320 which keep template borders and remain visible.
    # Data rows (17-305) also need border clearing — value=None alone is not enough
    # if the column isn't hidden properly.
    from openpyxl.styles import Border as _Border, PatternFill as _PFill
    _nb = _Border()
    _nf = _PFill()
    for pos in range(n + 1, 10):
        tc  = _t_data_col(pos)
        ltr = get_column_letter(tc)
        ws.column_dimensions[ltr].hidden = True
        ws.column_dimensions[ltr].width  = 0
        for r in range(1, 330):
            try:
                cell = ws.cell(row=r, column=tc)
                cell.value  = None
                cell.fill   = _copy_obj(_nf)
                cell.border = _copy_obj(_nb)
            except Exception: pass

    # ── Datos: timestamps y temperaturas filas 17-305 ───────────────────────────
    for i in range(rows):
        row = T_HR_DATA_ROW_START + i
        ws.cell(row=row, column=T_HR_COL_DATETIME).value = timestamps.iloc[i].to_pydatetime()
        for sensor in sensores:
            pos = sensor.posicion
            if pos < 1 or pos > n or sensor.datos is None or i >= len(sensor.datos):
                continue
            ws.cell(row=row, column=_t_data_col(pos)).value = _safe_temp(sensor.datos['temperatura'].iloc[i])

    # ── Metadata rows 10-15 para sensores > 9 (template ya tiene 1-9) ──────────
    for pos in range(10, n + 1):
        tc = _t_data_col(pos)
        pc = _prim_temp_col(pos)
        ws.cell(row=10, column=tc).value = f'=Primarios!{_cl(pc)}1'
        ws.cell(row=11, column=tc).value = f'=Primarios!{_cl(pc)}2'
        ws.cell(row=12, column=tc).value = f'=Primarios!{_cl(pc)}3'
        ws.cell(row=13, column=tc).value = f'=Primarios!{_cl(pc)}4'
        ws.cell(row=14, column=tc).value = f'=Primarios!{_cl(pc)}5'
        ws.cell(row=15, column=tc).value = f'=Primarios!{_cl(pc)}6'

    # ── Fórmulas de estadísticas por sensor (filas 306-309) ─────────────────────
    last_col = _cl(_t_data_col(n))
    for pos in range(1, n + 1):
        tc = _t_data_col(pos)
        c  = _cl(tc)
        ws.cell(row=306, column=tc).value = f'=MAX({c}{T_HR_DATA_ROW_START}:{c}305)'
        ws.cell(row=307, column=tc).value = f'=MIN({c}{T_HR_DATA_ROW_START}:{c}305)'
        ws.cell(row=308, column=tc).value = f'=ROUND(AVERAGE({c}{T_HR_DATA_ROW_START}:{c}305),2)'
        ws.cell(row=309, column=tc).value = f'=$D$311-{c}308'

    # ── Estadísticas globales (filas 311-313, siempre col D) ────────────────────
    first_col = _cl(_t_data_col(1))
    ws.cell(row=311, column=4).value = f'=ROUND(AVERAGE({first_col}308:{last_col}308),2)'
    ws.cell(row=312, column=4).value = f'=MIN({first_col}307:{last_col}307)'
    ws.cell(row=313, column=4).value = f'=MAX({first_col}306:{last_col}306)'

    # ── ANÁLISIS ESTADÍSTICO fila 321: actualizar rangos para N sensores ─────────
    ws.cell(row=321, column=2).value = f'=COUNT({first_col}308:{last_col}308)'
    ws.cell(row=321, column=3).value = f'=STDEV({first_col}308:{last_col}308)'
    # D321 = CONFIDENCE(0.05, C321, B321) — usa B y C, no necesita cambio de rango
    ws.cell(row=321, column=5).value = f'=ROUND(AVERAGE({first_col}308:{last_col}308),2)'
    # F321 = E321-D321, G321 = E321+D321 — sin rango, no cambian


# ─────────────────────────────────────────────────────────────────
#  HR SHEET (humidity + GHR chart data)
# ─────────────────────────────────────────────────────────────────

def _llenar_hoja_HR(wb, sensores: List[Sensor], timestamps: pd.Series, n: int) -> None:
    ws = wb['HR']
    rows = min(len(timestamps), MAX_PRIMARIOS)

    # ── Estilos: copiar columna sensor-9 a columnas nuevas (n > 9) ─────────────
    for pos in range(10, n + 1):
        _estilizar_col_sensor(ws, pos, ref_pos=9)

    # ── Limpiar y ocultar columnas no usadas cuando n < 9 ──────────────────────
    # Same comprehensive clear as _llenar_hoja_T — rows 1-330 covers header,
    # data, stats, and all template gap rows (310, 314-320, etc.).
    from openpyxl.styles import Border as _Border2, PatternFill as _PFill2
    _nb2 = _Border2()
    _nf2 = _PFill2()
    for pos in range(n + 1, 10):
        tc  = _t_data_col(pos)
        ltr = get_column_letter(tc)
        ws.column_dimensions[ltr].hidden = True
        ws.column_dimensions[ltr].width  = 0
        for r in range(1, 330):
            try:
                cell = ws.cell(row=r, column=tc)
                cell.value  = None
                cell.fill   = _copy_obj(_nf2)
                cell.border = _copy_obj(_nb2)
            except Exception: pass

    # ── Datos: timestamps y humedad filas 17-305 ────────────────────────────────
    for i in range(rows):
        row = T_HR_DATA_ROW_START + i
        ws.cell(row=row, column=T_HR_COL_DATETIME).value = timestamps.iloc[i].to_pydatetime()
        for sensor in sensores:
            pos = sensor.posicion
            if pos < 1 or pos > n or sensor.datos is None or i >= len(sensor.datos):
                continue
            ws.cell(row=row, column=_t_data_col(pos)).value = _safe_hum(sensor.datos['humedad'].iloc[i])

    # ── Metadata rows 10-15: reescribir SIEMPRE para todas las posiciones 1..n ────
    # The N=9 template has hardcoded refs to Primarios humidity cols M-U.
    # _prim_hum_col(pos, n) = n+3+pos — this shifts with N, so for N≠9 the template
    # refs are wrong. Rewrite every active position, not only pos > 9.
    for pos in range(1, n + 1):
        tc = _t_data_col(pos)
        hc = _prim_hum_col(pos, n)
        ws.cell(row=10, column=tc).value = f'=Primarios!{_cl(hc)}1'
        ws.cell(row=11, column=tc).value = f'=Primarios!{_cl(hc)}2'
        ws.cell(row=12, column=tc).value = f'=Primarios!{_cl(hc)}3'
        ws.cell(row=13, column=tc).value = f'=Primarios!{_cl(hc)}4'
        ws.cell(row=14, column=tc).value = f'=Primarios!{_cl(hc)}5'
        ws.cell(row=15, column=tc).value = f'=Primarios!{_cl(hc)}6'

    # ── Fórmulas de estadísticas por sensor (filas 306-309) ─────────────────────
    last_col  = _cl(_t_data_col(n))
    first_col = _cl(_t_data_col(1))
    for pos in range(1, n + 1):
        tc = _t_data_col(pos)
        c  = _cl(tc)
        ws.cell(row=306, column=tc).value = f'=MAX({c}{T_HR_DATA_ROW_START}:{c}305)'
        ws.cell(row=307, column=tc).value = f'=MIN({c}{T_HR_DATA_ROW_START}:{c}305)'
        ws.cell(row=308, column=tc).value = f'=ROUND(AVERAGE({c}{T_HR_DATA_ROW_START}:{c}305),2)'
        ws.cell(row=309, column=tc).value = f'=$D$311-{c}308'

    # ── Estadísticas globales (filas 311-313, siempre col D) ────────────────────
    ws.cell(row=311, column=4).value = f'=ROUND(AVERAGE({first_col}308:{last_col}308),2)'
    ws.cell(row=312, column=4).value = f'=MIN({first_col}307:{last_col}307)'
    ws.cell(row=313, column=4).value = f'=MAX({first_col}306:{last_col}306)'

    # ── ANÁLISIS ESTADÍSTICO fila 321: actualizar rangos para N sensores ─────────
    ws.cell(row=321, column=2).value = f'=COUNT({first_col}308:{last_col}308)'
    ws.cell(row=321, column=3).value = f'=STDEV({first_col}308:{last_col}308)'
    ws.cell(row=321, column=5).value = f'=ROUND(AVERAGE({first_col}308:{last_col}308),2)'


# ─────────────────────────────────────────────────────────────────
#  FALLAS
# ─────────────────────────────────────────────────────────────────

def _escribir_nombres_fallas(wb, sheet_name: str, tipo_prueba: str) -> None:
    ws = wb[sheet_name]
    nombre_fase1 = 'Corte de energía' if tipo_prueba == 'PO' else 'Apertura de puerta'
    ws.cell(row=2, column=7).value = nombre_fase1
    ws.cell(row=3, column=7).value = 'Recuperación'


def _llenar_fallas(wb, sheet_name: str, df: pd.DataFrame, es_temp: bool) -> None:
    ws = wb[sheet_name]
    for i, (_, fila) in enumerate(df.iterrows()):
        if i >= MAX_FALLAS:
            break
        row  = FALLAS_ROW_START + i
        cell = ws.cell(row=row, column=2)
        cell.value = fila['timestamp']
        cell.number_format = 'DD/MM/YYYY h:mm am/pm'
        ws.cell(row=row, column=3).value = (_safe_temp if es_temp else _safe_hum)(fila['valor'])


# ─────────────────────────────────────────────────────────────────
#  ANÁLISIS
# ─────────────────────────────────────────────────────────────────

def _llenar_analisis(wb, config: ProyectoConfig, n: int, sensores: List[Sensor]) -> None:
    nombre = _resolver_sheet(wb, ['Análisis', 'Analisis', 'An\u00e1lisis'])
    if not nombre:
        return
    ws = wb[nombre]

    # Basic config cells
    ws['C2'] = config.empresa
    ws['C3'] = config.marca_equipo
    ws['C4'] = config.ubicacion
    ws['C5'] = config.codigo_equipo
    ws['C6'] = config.setpoint_temp
    ws['C7'] = config.setpoint_hum

    # ── Tabla 1 header: "Recolectores (+)" row (13) and sensor header rows (14-15) ──
    # Template: C13:L13 merged (sensors 1-9 in cols D=4..L=12), rows 14-15 have blue fill.
    # We must: extend/shrink the merge to col 3+n, copy styling, clear unused cols.
    _T1_HDR    = 13  # "Recolectores (+)" row
    _T1_POS    = 14  # "Posición X" row
    _T1_SER    = 15  # serial row
    _T1_NOTE   = 16  # note row (only pos1 has "(++)")
    _T1_S_COL  = 4   # first sensor col (D = pos 1)
    _t1_last   = 3 + n  # last sensor col (L=12 for n=9, M=13 for n=10)

    # Capture reference style from an existing sensor header cell (pos 1 = D14)
    _t1_ref = ws.cell(row=_T1_POS, column=_T1_S_COL)
    _t1_fill = _copy_obj(_t1_ref.fill)
    _t1_font = _copy_obj(_t1_ref.font)

    # Unmerge the existing "Recolectores (+)" merged range on row 13
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == _T1_HDR and mr.max_row == _T1_HDR and mr.min_col == 3:
            ws.unmerge_cells(str(mr))
            break

    # Re-apply fill to all secondary cells in the row-13 merged range (needed before re-merge)
    for col in range(3, _t1_last + 1):
        cell = ws.cell(row=_T1_HDR, column=col)
        if col > 3:
            cell.fill = _copy_obj(_t1_fill)

    # Re-merge row 13 to exactly cover active sensors
    ws.merge_cells(start_row=_T1_HDR, start_column=3,
                   end_row=_T1_HDR, end_column=_t1_last)

    # For N<9: old secondary cells of the N=9 merge (cols _t1_last+1 .. 12) keep their
    # fill after unmerge. Clear both fill and border so they don't appear as ghost cells.
    _no_fill_a  = PatternFill()
    _no_bord_a  = openpyxl.styles.Border()
    if n < 9:
        for col in range(_t1_last + 1, 13):   # 13 = col L + 1 = N=9 template max + 1
            try:
                c = ws.cell(row=_T1_HDR, column=col)
                c.fill   = _copy_obj(_no_fill_a)
                c.border = _copy_obj(_no_bord_a)
            except Exception:
                pass

    # Clear and unmerge unused sensor cols (n < 9) in rows 14-16
    for pos in range(n + 1, 10):
        col = 3 + pos
        for r in (_T1_POS, _T1_SER, _T1_NOTE):
            try:
                cell = ws.cell(row=r, column=col)
                cell.value  = None
                cell.fill   = _copy_obj(_no_fill_a)
                cell.border = _copy_obj(_no_bord_a)
            except Exception:
                pass
        # Unmerge serial cells (E15:E16 … L15:L16) for unused positions
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row == _T1_SER and mr.max_row == _T1_NOTE and mr.min_col == col:
                ws.unmerge_cells(str(mr))

    # Reference style for serial rows: use position-2 (col E = _T1_S_COL+1)
    # which is a merged E15:E16 cell with the correct blue fill/border/font.
    _t1_ser_ref = ws.cell(row=_T1_SER, column=_T1_S_COL + 1)

    # Write and style new sensor cols (n > 9) in rows 14-15, add row 15-16 merge
    for pos in range(10, n + 1):
        col  = 3 + pos
        pc_t = _prim_temp_col(pos)
        # Row 14: position label with blue fill + bold font
        c14 = ws.cell(row=_T1_POS, column=col)
        c14.value = f'=Primarios!{_cl(pc_t)}5'
        c14.fill  = _copy_obj(_t1_fill)
        c14.font  = _copy_obj(_t1_font)
        # Row 15: serial — copy full style from reference pos-2 cell
        c15 = ws.cell(row=_T1_SER, column=col)
        c15.value = f'=Primarios!{_cl(pc_t)}3'
        _copy_cell_style(_t1_ser_ref, c15)
        # Merge rows 15-16 for the serial cell (same as template pos 2-9)
        try:
            ws.merge_cells(start_row=_T1_SER, start_column=col,
                           end_row=_T1_NOTE, end_column=col)
        except Exception:
            pass

    # Row positions.
    # Template is designed for N=9 with row_global=32. For N<9 no rows are
    # inserted so every row stays at the template position. For N>9 we inserted
    # (n-9) extra rows so every row shifts by that amount.
    # WRONG: row_global = 23 + n  (gives 29 for N=6 — writes formulas 3 rows
    #        above the template labels, producing "random" misplaced values)
    # CORRECT: base is always 32; only add the extra rows actually inserted.
    extra       = max(0, n - 9)
    row_global  = 32 + extra       # global avg summary row
    row_t_check = row_global + 7   # Trecolector temp  (row 39 for N≤9)
    row_h_check = row_global + 10  # Trecolector HR    (row 42 for N≤9)
    row_t5_t    = row_global + 17  # Tabla 5 temp data (row 49 for N≤9)
    row_t5_h    = row_global + 20  # Tabla 5 HR data   (row 52 for N≤9)
    row_t6_id   = row_global + 25  # Tabla 6 identification (row 57 for N≤9)
    row_t6_ser  = row_global + 26  # Tabla 6 serial    (row 58 for N≤9)
    row_t6_ta   = row_global + 27  # Tabla 6 temp avg  (row 59 for N≤9)
    row_t6_ha   = row_global + 28  # Tabla 6 HR avg    (row 60 for N≤9)

    # Clear AND HIDE unused sensor rows (only needed when n < 9).
    # Rows 23+n .. 31 are template rows for sensors n+1 .. 9.
    # Clearing value+fill+border removes styling but leaves empty white rows that
    # visually split the table from the global-avg row. Hiding removes the gap entirely.
    for r in range(23 + n, 32):
        ws.row_dimensions[r].hidden = True
        for col in range(1, 13):
            try:
                cell = ws.cell(row=r, column=col)
                cell.value  = None
                cell.fill   = _copy_obj(_no_fill_a)
                cell.border = _copy_obj(_no_bord_a)
            except Exception:
                pass

    # Extend conditional formatting to cover all N sensor rows (template hard-codes row 31).
    # Bug fix: cf_obj.sqref is a MultiCellRange object, not a string — must use str() before
    # calling .replace(), otherwise the call fails silently inside the try/except.
    _last_sensor_row = 22 + n
    try:
        from openpyxl.formatting.formatting import ConditionalFormatting as _CFClass
        _cf_updates = {
            'D23:E31': f'D23:E{_last_sensor_row}',
            'J23:K31': f'J23:K{_last_sensor_row}',
        }
        _cf = ws.conditional_formatting
        _new_cf_rules = {}
        for cf_obj, rules in list(_cf._cf_rules.items()):
            sqref_str = str(cf_obj.sqref) if hasattr(cf_obj, 'sqref') else str(cf_obj)
            new_sqref = sqref_str
            for old_range, new_range in _cf_updates.items():
                new_sqref = new_sqref.replace(old_range, new_range)
            if new_sqref != sqref_str:
                new_cf_obj = _CFClass(sqref=new_sqref)
                _new_cf_rules[new_cf_obj] = rules
            else:
                _new_cf_rules[cf_obj] = rules
        _cf._cf_rules = _new_cf_rules
    except Exception:
        pass

    # Add NO→red rule for the full sensor range.
    # The template only has SI→green; "NO"/"NO1"/"NO2" cells must be explicitly red.
    # 8-char ARGB: 'FF' prefix = fully opaque.
    _red_fill   = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    _green_fill = PatternFill(start_color='FF00B050', end_color='FF00B050', fill_type='solid')
    ws.conditional_formatting.add(
        f'D23:E{_last_sensor_row}',
        CellIsRule(operator='equal', formula=['"SI"'], fill=_green_fill))
    ws.conditional_formatting.add(
        f'J23:K{_last_sensor_row}',
        CellIsRule(operator='equal', formula=['"SI"'], fill=_green_fill))
    ws.conditional_formatting.add(
        f'D23:E{_last_sensor_row}',
        CellIsRule(operator='notEqual', formula=['"SI"'], fill=_red_fill))
    ws.conditional_formatting.add(
        f'J23:K{_last_sensor_row}',
        CellIsRule(operator='notEqual', formula=['"SI"'], fill=_red_fill))

    # Font for new sensor rows (pos > 9) — explicit Arial 8pt black to avoid theme-color drift
    _row_font = openpyxl.styles.Font(name='Arial', size=8, bold=False)

    # Write N sensor rows (23 to 22+n)
    for pos in range(1, n + 1):
        r   = 22 + pos
        tc  = _t_data_col(pos)
        col = _cl(tc)   # D, E, F, ...

        ws.cell(row=r, column=1).value  = f'={_cl(tc)}15'
        ws.cell(row=r, column=2).value  = f'=T!{col}308'
        ws.cell(row=r, column=3).value  = f'=B${row_global}-B{r}'
        ws.cell(row=r, column=4).value  = f'=IF(ABS(C{r})<2,"SI","NO")'
        ws.cell(row=r, column=5).value  = f'=IF(B{r}>($C$6-2),(IF(B{r}<($C$6+2),"SI","NO2")),"NO1")'
        ws.cell(row=r, column=7).value  = f'=A{r}'
        ws.cell(row=r, column=8).value  = f'=HR!{col}308'
        ws.cell(row=r, column=9).value  = f'=H${row_global}-H{r}'
        ws.cell(row=r, column=10).value = f'=IF(ABS(I{r})<5,"SI","NO")'
        ws.cell(row=r, column=11).value = f'=IF(H{r}>($C$7-5),(IF(H{r}<($C$7+5),"SI","NO2")),"NO1")'

        # Apply explicit font to new rows so they match the template sensor rows
        if pos > 9:
            for c in range(1, 12):
                try: ws.cell(row=r, column=c).font = _copy_obj(_row_font)
                except Exception: pass

    # Global avg row
    ws.cell(row=row_global, column=2).value = '=T!D311'
    ws.cell(row=row_global, column=8).value = '=HR!D311'

    # Trecolector temp (row_t_check)
    ws.cell(row=row_t_check, column=1).value = config.setpoint_temp
    ws.cell(row=row_t_check, column=2).value = '=T!D311'
    ws.cell(row=row_t_check, column=3).value = '=T!D312'
    ws.cell(row=row_t_check, column=4).value = '=T!D313'
    ws.cell(row=row_t_check, column=5).value = f'=A{row_t_check}-C{row_t_check}'
    ws.cell(row=row_t_check, column=6).value = f'=D{row_t_check}-A{row_t_check}'
    ws.cell(row=row_t_check, column=7).value = f'=IF(ABS(E{row_t_check})<2,IF(ABS(F{row_t_check}<2),"SI","NO"),"NO")'

    # Trecolector HR (row_h_check)
    ws.cell(row=row_h_check, column=1).value = config.setpoint_hum
    ws.cell(row=row_h_check, column=2).value = '=HR!D311'
    ws.cell(row=row_h_check, column=3).value = '=HR!D312'
    ws.cell(row=row_h_check, column=4).value = '=HR!D313'
    ws.cell(row=row_h_check, column=5).value = f'=A{row_h_check}-C{row_h_check}'
    ws.cell(row=row_h_check, column=6).value = f'=D{row_h_check}-A{row_h_check}'
    ws.cell(row=row_h_check, column=7).value = f'=IF(ABS(E{row_h_check})<5,IF(ABS(F{row_h_check}<5),"SI","NO"),"NO")'

    # Tabla 5 — temp
    ws.cell(row=row_t5_t, column=1).value = f'=B23'
    ws.cell(row=row_t5_t, column=3).value = f'=B{row_t5_t}-A{row_t5_t}'
    ws.cell(row=row_t5_t, column=4).value = f'=IF(ABS(C{row_t5_t})<1,"SI","NO")'
    ws.cell(row=row_t5_t, column=2).value = round(float(config.lectura_equipo_temp), 2)

    # Tabla 5 — HR
    ws.cell(row=row_t5_h, column=1).value = f'=H23'
    ws.cell(row=row_t5_h, column=3).value = f'=B{row_t5_h}-A{row_t5_h}'
    ws.cell(row=row_t5_h, column=4).value = f'=IF(ABS(C{row_t5_h})<1,"SI","NO")'
    ws.cell(row=row_t5_h, column=2).value = round(float(config.lectura_equipo_hum), 1)

    # Tabla 6 — localización de puntos críticos
    # Template uses cols B-J (pos+1) for sensors 1-9; extend naturally for N>9.
    # Reference column is J (pos9) — copy its borders/style to new sensor cols.
    _T6_REF_COL = 10  # col J = pos 9 in template
    for pos in range(1, n + 1):
        tc     = _t_data_col(pos)
        col    = _cl(tc)
        t6_col = pos + 1  # B=2 for pos1, K=11 for pos10, L=12 for pos11...
        ws.cell(row=row_t6_id,  column=t6_col).value = f'={col}14'
        ws.cell(row=row_t6_ser, column=t6_col).value = f'={col}15'
        ws.cell(row=row_t6_ta,  column=t6_col).value = f'=T!{col}308'
        ws.cell(row=row_t6_ha,  column=t6_col).value = f'=HR!{col}308'
        # For new sensor columns (beyond original 9), copy border+font style from reference col J
        if pos > 9:
            for row in (row_t6_id, row_t6_ser, row_t6_ta, row_t6_ha):
                _copy_cell_style(ws.cell(row=row, column=_T6_REF_COL),
                                 ws.cell(row=row, column=t6_col))

    # Clear unused Tabla 6 sensor columns when n < 9 (template has cols B-J for 9 sensors)
    for pos in range(n + 1, 10):
        t6_col = pos + 1
        for row in (row_t6_id, row_t6_ser, row_t6_ta, row_t6_ha):
            try:
                ws.cell(row=row, column=t6_col).value = None
                ws.cell(row=row, column=t6_col).border = openpyxl.styles.Border()
            except Exception: pass

    # Set Tabla 6 row heights (template has 20.4 for temp row, 26.25 for HR row)
    ws.row_dimensions[row_t6_ta].height = 20.4
    ws.row_dimensions[row_t6_ha].height = 26.25

    # Ensure new sensor columns have correct width (13.0 to match template cols D-J)
    for pos in range(10, n + 1):
        ltr = _cl(pos + 1)
        ws.column_dimensions[ltr].width = 13.0

    # Update MAX/MIN promedio formulas to cover all N sensors
    last_t6 = _cl(n + 1)
    ws.cell(row=row_t6_ta + 4, column=2).value = f'=MAX(B{row_t6_ta}:{last_t6}{row_t6_ta})'
    ws.cell(row=row_t6_ta + 6, column=2).value = f'=MIN(B{row_t6_ta}:{last_t6}{row_t6_ta})'
    ws.cell(row=row_t6_ha + 8, column=2).value = f'=MAX(B{row_t6_ha}:{last_t6}{row_t6_ha})'
    ws.cell(row=row_t6_ha + 10, column=2).value = f'=MIN(B{row_t6_ha}:{last_t6}{row_t6_ha})'

    # Color extremes in Tabla 6:
    #   Temp más caliente → rojo  | Temp más frío   → azul claro
    #   HR   más húmedo   → verde | HR   menos húmedo → amarillo
    # openpyxl no evalúa fórmulas, así que calculamos los promedios directamente
    # desde los datos de cada sensor.
    _temp_avgs: dict[int, float] = {}
    _hum_avgs:  dict[int, float] = {}
    for s in sensores:
        if s.datos is None or len(s.datos) == 0:
            continue
        if 'temperatura' in s.datos.columns:
            vals = s.datos['temperatura'].dropna()
            if len(vals) > 0:
                _temp_avgs[s.posicion] = float(vals.mean())
        if 'humedad' in s.datos.columns:
            vals = s.datos['humedad'].dropna()
            if len(vals) > 0:
                _hum_avgs[s.posicion] = float(vals.mean())

    _fill_hot   = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    _fill_cold  = PatternFill(start_color='FF00B0F0', end_color='FF00B0F0', fill_type='solid')
    _fill_humid = PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid')
    _fill_dry   = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    if _temp_avgs:
        _max_t_pos = max(_temp_avgs, key=_temp_avgs.get)
        _min_t_pos = min(_temp_avgs, key=_temp_avgs.get)
        try: ws.cell(row=row_t6_ta, column=_max_t_pos + 1).fill = _fill_hot
        except Exception: pass
        try: ws.cell(row=row_t6_ta, column=_min_t_pos + 1).fill = _fill_cold
        except Exception: pass

    if _hum_avgs:
        _max_h_pos = max(_hum_avgs, key=_hum_avgs.get)
        _min_h_pos = min(_hum_avgs, key=_hum_avgs.get)
        try: ws.cell(row=row_t6_ha, column=_max_h_pos + 1).fill = _fill_humid
        except Exception: pass
        try: ws.cell(row=row_t6_ha, column=_min_h_pos + 1).fill = _fill_dry
        except Exception: pass


# ─────────────────────────────────────────────────────────────────
#  CHART XML PATCHING — N series for GT and GHR
# ─────────────────────────────────────────────────────────────────

def _serie_xml(idx: int, sheet: str, col: str) -> str:
    spPr = _CHART_LINE_STYLES[idx % len(_CHART_LINE_STYLES)]
    return (
        f'<c:ser>'
        f'<c:idx val="{idx}"/><c:order val="{idx}"/>'
        f'<c:tx><c:strRef><c:f>{sheet}!${col}$14</c:f>'
        f'<c:strCache><c:ptCount val="1"/>'
        f'<c:pt idx="0"><c:v>Posici\u00f3n {idx+1}</c:v></c:pt>'
        f'</c:strCache></c:strRef></c:tx>'
        f'<c:spPr>{spPr}</c:spPr>'
        f'<c:marker><c:symbol val="none"/></c:marker>'
        f'<c:xVal><c:numRef>'
        f'<c:f>{sheet}!$C$17:$C$305</c:f>'
        f'<c:numCache>'
        f'<c:formatCode>dd/mm/yyyy\\ hh:mm\\ AM/PM</c:formatCode>'
        f'<c:ptCount val="0"/>'
        f'</c:numCache></c:numRef></c:xVal>'
        f'<c:val><c:numRef>'
        f'<c:f>{sheet}!${col}$17:${col}$305</c:f>'
        f'<c:numCache><c:formatCode>0.00</c:formatCode>'
        f'<c:ptCount val="0"/>'
        f'</c:numCache></c:numRef></c:val>'
        f'</c:ser>'
    )


def _patch_chart_n_sensores(ruta_salida: str, n: int) -> None:
    """Replace series in GT (chart1) and GHR (chart3) for exactly N sensors."""
    contenido: dict[str, bytes] = {}
    with zipfile.ZipFile(ruta_salida, 'r') as zo:
        for name in zo.namelist():
            contenido[name] = zo.read(name)

    for chart_file, sheet in [('xl/charts/chart1.xml', 'T'),
                               ('xl/charts/chart3.xml', 'HR')]:
        if chart_file not in contenido:
            continue
        xml = contenido[chart_file].decode('utf-8', errors='replace')

        # Build new series block for N sensors
        new_series = ''.join(
            _serie_xml(i, sheet, _cl(_t_data_col(i + 1)))
            for i in range(n)
        )

        # Replace series block IN-PLACE using a single greedy match.
        # Non-greedy .*? would remove each <c:ser> individually and then
        # append new series before </c:scatterChart>, placing them AFTER
        # <c:axId> elements and breaking the chart layout/margins.
        # Greedy .* captures from the first <c:ser> to the LAST </c:ser>
        # and replaces the whole block at its original position, keeping
        # <c:axId> and all other chart elements where they belong.
        if '<c:ser>' in xml:
            xml = re.sub(r'<c:ser>.*</c:ser>', new_series, xml,
                         count=1, flags=re.DOTALL)
        else:
            # No existing series — insert before the chart closing tag
            for anchor in ['</c:scatterChart>', '</c:lineChart>', '</c:barChart>']:
                if anchor in xml:
                    xml = xml.replace(anchor, new_series + anchor, 1)
                    break

        contenido[chart_file] = xml.encode('utf-8')

    tmp = ruta_salida + '.chart.tmp'
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zn:
        for name, data in contenido.items():
            zn.writestr(name, data)
    os.replace(tmp, ruta_salida)


# ─────────────────────────────────────────────────────────────────
#  CUADROS AZULES (userShapes ZIP patching)
# ─────────────────────────────────────────────────────────────────

def _fmt_dt(dt) -> str:
    ap = 'a.m.' if dt.hour < 12 else 'p.m.'
    h  = dt.hour % 12 or 12
    return f'{h:02d}:{dt.minute:02d}:{dt.second:02d} {ap} / {dt.strftime("%Y-%m-%d")}'


def _fmt_rango(t1, t2) -> str:
    def _t(dt):
        ap = 'a.m.' if dt.hour < 12 else 'p.m.'
        h  = dt.hour % 12 or 12
        return f'{h:02d}:{dt.minute:02d}:{dt.second:02d} {ap}'
    return f'{_t(t1)} a {_t(t2)} / {t2.strftime("%Y-%m-%d")}'


def _reemplazar_texto(xml: str, old: str, new: str) -> str:
    old_esc = old.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    new_esc = new.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    return xml.replace(old_esc, new_esc)


def _inyectar_cuadros_info(ruta_plantilla: str, ruta_salida: str, config) -> None:
    if config.tipo_prueba == 'PO':
        duracion_fase1 = timedelta(minutes=30)
        duracion_total = timedelta(minutes=60)
        label_fase1    = 'Hora y fecha de corte de energ\u00eda:\t'
    else:
        duracion_fase1 = timedelta(minutes=5)
        duracion_total = timedelta(minutes=35)
        label_fase1    = 'Hora y fecha de apertura de puerta:\t'

    fin_fase1   = config.inicio_falla + duracion_fase1
    fin_total   = config.inicio_falla + duracion_total
    fin_24h     = config.inicio_24h + timedelta(hours=24)
    ensayo_base = config.ensayo or ''

    with zipfile.ZipFile(ruta_plantilla, 'r') as zt:
        template_drawings = {k: zt.read(k).decode('utf-8', errors='replace')
                             for k in _USERSHAPES_MAP}

    def _fill_24h(xml: str, _var: str) -> str:
        xml = _reemplazar_texto(xml, 'XXXXXXXXX', config.ubicacion)
        xml = _reemplazar_texto(xml,
                                'Hora y fecha de Inicio:\t00:00:00 pm / AAAA-MM-DD',
                                f'Hora y fecha de Inicio:\t{_fmt_dt(config.inicio_24h)}')
        xml = _reemplazar_texto(xml,
                                'Hora y fecha Final:\t\t00:00:00 pm / AAAA-MM-DD',
                                f'Hora y fecha Final:\t\t{_fmt_dt(fin_24h)}')
        xml = _reemplazar_texto(xml, 'XXXXXXXX POI', f'{ensayo_base} POI')
        return xml

    def _fill_fallas(xml: str, _var: str) -> str:
        xml = _reemplazar_texto(xml, 'Ubicaci\u00f3n:\t\t',
                                f'Ubicaci\u00f3n:\t\t{config.ubicacion}')
        xml = _reemplazar_texto(xml, 'Hora y fecha de corte de energ\u00eda:\t',
                                f'{label_fase1}{_fmt_rango(config.inicio_falla, fin_fase1)}')
        xml = _reemplazar_texto(xml, 'Hora y fecha de recuperaci\u00f3n:\t\t',
                                f'Hora y fecha de recuperaci\u00f3n:\t\t{_fmt_rango(fin_fase1, fin_total)}')
        xml = _reemplazar_texto(xml, 'Ensayo:\t\t', f'Ensayo:\t\t{ensayo_base} PPI')
        return xml

    filled = {
        'xl/drawings/drawing2.xml': _fill_24h(template_drawings['xl/drawings/drawing2.xml'], 'T'),
        'xl/drawings/drawing6.xml': _fill_24h(template_drawings['xl/drawings/drawing6.xml'], 'HR'),
        'xl/drawings/drawing4.xml': _fill_fallas(template_drawings['xl/drawings/drawing4.xml'], 'T'),
        'xl/drawings/drawing8.xml': _fill_fallas(template_drawings['xl/drawings/drawing8.xml'], 'HR'),
    }

    chart_rels = {
        'xl/charts/_rels/chart1.xml.rels': '../drawings/us_gt.xml',
        'xl/charts/_rels/chart2.xml.rels': '../drawings/us_fallastem.xml',
        'xl/charts/_rels/chart3.xml.rels': '../drawings/us_ghr.xml',
        'xl/charts/_rels/chart4.xml.rels': '../drawings/us_fallashr.xml',
    }
    chart_us_target = {
        'xl/charts/chart1.xml': '../drawings/us_gt.xml',
        'xl/charts/chart2.xml': '../drawings/us_fallastem.xml',
        'xl/charts/chart3.xml': '../drawings/us_ghr.xml',
        'xl/charts/chart4.xml': '../drawings/us_fallashr.xml',
    }

    contenido: dict[str, bytes] = {}
    with zipfile.ZipFile(ruta_salida, 'r') as zo:
        for name in zo.namelist():
            contenido[name] = zo.read(name)

    for src, dst in _USERSHAPES_MAP.items():
        contenido[dst] = filled[src].encode('utf-8')

    for rels_path, target in chart_rels.items():
        contenido[rels_path] = _CHART_RELS_TEMPLATE.format(target=target).encode('utf-8')

    _R_NS = 'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
    for chart_xml, _ in chart_us_target.items():
        if chart_xml in contenido:
            xml = contenido[chart_xml].decode('utf-8')
            if 'userShapes' not in xml:
                if '</c:chartSpace>' in xml:
                    if 'xmlns:r=' not in xml:
                        xml = xml.replace('<c:chartSpace ', f'<c:chartSpace {_R_NS} ', 1)
                    xml = xml.replace('</c:chartSpace>',
                                      '<c:userShapes r:id="rId3"/></c:chartSpace>')
                elif '</chartSpace>' in xml:
                    if 'xmlns:r=' not in xml:
                        xml = xml.replace('<chartSpace ', f'<chartSpace {_R_NS} ', 1)
                    xml = xml.replace('</chartSpace>',
                                      '<userShapes r:id="rId3"/></chartSpace>')
                contenido[chart_xml] = xml.encode('utf-8')

    ct_xml = contenido['[Content_Types].xml'].decode('utf-8')
    ct_type = 'application/vnd.openxmlformats-officedocument.drawingml.chartshapes+xml'
    for dst in _USERSHAPES_MAP.values():
        part  = '/' + dst
        entry = f'<Override PartName="{part}" ContentType="{ct_type}"/>'
        if part not in ct_xml:
            ct_xml = ct_xml.replace('</Types>', entry + '</Types>')
    contenido['[Content_Types].xml'] = ct_xml.encode('utf-8')

    tmp = ruta_salida + '.info.tmp'
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zn:
        for name, data in contenido.items():
            zn.writestr(name, data)
    for _attempt in range(6):
        try:
            os.replace(tmp, ruta_salida)
            break
        except PermissionError:
            if _attempt == 5:
                raise
            time.sleep(0.4)


# ─────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────

def _safe_temp(val) -> float | None:
    try:
        v = float(val)
        return round(v, 2) if not pd.isna(v) else None
    except (TypeError, ValueError):
        return None


def _safe_hum(val) -> float | None:
    try:
        v = float(val)
        return round(v, 1) if not pd.isna(v) else None
    except (TypeError, ValueError):
        return None
