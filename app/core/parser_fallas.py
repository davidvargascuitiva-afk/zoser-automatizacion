import openpyxl
import pandas as pd
from datetime import datetime
from typing import Optional, Tuple


def _extraer_datetime(fecha_val, tiempo_val) -> Optional[datetime]:
    if fecha_val is None:
        return None
    if isinstance(fecha_val, datetime):
        if fecha_val.hour != 0 or fecha_val.minute != 0:
            return fecha_val
        if isinstance(tiempo_val, datetime):
            return fecha_val.replace(
                hour=tiempo_val.hour,
                minute=tiempo_val.minute,
                second=tiempo_val.second,
            )
        return fecha_val
    return None


def obtener_rango_fechas(filepath: str):
    """Returns (min_datetime, max_datetime) from the fallas file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    timestamps = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        fecha_val  = row[0]
        tiempo_val = row[1] if len(row) > 1 else None
        if fecha_val is None:
            continue
        ts = _extraer_datetime(fecha_val, tiempo_val)
        if ts is not None:
            timestamps.append(ts)
    wb.close()
    if not timestamps:
        return None, None
    return min(timestamps), max(timestamps)


def cargar_datos_fallas(
    filepath: str,
    tipo_prueba: str,
    inicio_falla: datetime,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns (df_temp, df_hum) each with columns [timestamp, valor].
    PO: 60 records (30 min corte + 30 min recovery @ 1/min)
    PP: 35 records (5 min puerta + 30 min recovery @ 1/min)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        fecha_val  = row[0]
        tiempo_val = row[1]
        if fecha_val is None:
            continue
        ts = _extraer_datetime(fecha_val, tiempo_val)
        if ts is None:
            continue
        temp = row[2] if len(row) > 2 else None
        hum  = row[3] if len(row) > 3 else None
        rows.append({'timestamp': ts, 'temperatura': temp, 'humedad': hum})

    wb.close()

    empty = pd.DataFrame(columns=['timestamp', 'valor'])
    if not rows:
        return empty.copy(), empty.copy()

    df = pd.DataFrame(rows)
    df['temperatura'] = pd.to_numeric(df['temperatura'], errors='coerce')
    df['humedad']     = pd.to_numeric(df['humedad'],     errors='coerce')

    df = df[df['timestamp'] >= inicio_falla].reset_index(drop=True)

    num_registros = 60 if tipo_prueba == "PO" else 35
    df = df.head(num_registros)

    df_temp = df[['timestamp', 'temperatura']].rename(columns={'temperatura': 'valor'})
    df_hum  = df[['timestamp', 'humedad']].rename(columns={'humedad': 'valor'})

    return df_temp, df_hum
