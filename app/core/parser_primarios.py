import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime, date, timedelta
from typing import List, Optional, Dict
from app.models.sensor import Sensor


def _extraer_datetime(fecha_val, tiempo_val) -> Optional[datetime]:
    """Combines separate date and time cells from MadgeTech export."""
    if fecha_val is None:
        return None

    if isinstance(fecha_val, datetime):
        if fecha_val.hour != 0 or fecha_val.minute != 0 or fecha_val.second != 0:
            return fecha_val
        if isinstance(tiempo_val, datetime):
            return fecha_val.replace(
                hour=tiempo_val.hour,
                minute=tiempo_val.minute,
                second=tiempo_val.second,
            )
        return fecha_val

    if isinstance(fecha_val, date):
        if isinstance(tiempo_val, datetime):
            return datetime.combine(fecha_val, tiempo_val.time())

    return None


def detectar_sensores(filepath: str) -> List[Sensor]:
    """Reads MadgeTech MultiChannel file and returns detected sensors in source order."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    sensor_map: Dict[str, dict] = {}

    for col_idx in range(3, ws.max_column + 1):
        nombre = ws.cell(row=1, column=col_idx).value
        desc   = ws.cell(row=2, column=col_idx).value
        serial = ws.cell(row=3, column=col_idx).value
        header = ws.cell(row=7, column=col_idx).value

        if serial is None or header is None:
            continue

        serial     = str(serial).strip()
        header_low = str(header).lower()

        if serial not in sensor_map:
            sensor_map[serial] = {
                'nombre':      str(nombre).strip() if nombre else '',
                'descripcion': str(desc).strip()   if desc   else '',
                'col_temp':    None,
                'col_hum':     None,
            }

        if any(k in header_low for k in ('temperatura', '\u00b0c', 'temperature')):
            sensor_map[serial]['col_temp'] = col_idx
        elif any(k in header_low for k in ('humedad', 'rh', 'humidity')):
            sensor_map[serial]['col_hum'] = col_idx

    wb.close()

    sensores = []
    for pos, (serial, info) in enumerate(sensor_map.items(), start=1):
        sensores.append(Sensor(
            serial=serial,
            nombre=info['nombre'],
            descripcion=info['descripcion'],
            posicion=pos,
            tiene_temperatura=info['col_temp'] is not None,
            tiene_humedad=info['col_hum']  is not None,
            col_idx_temp=info['col_temp'],
            col_idx_hum=info['col_hum'],
        ))

    return sensores


def obtener_rango_fechas(filepath: str):
    """Returns (min_datetime, max_datetime) from the primarios file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    timestamps = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        fecha_val  = row[0]
        tiempo_val = row[1] if len(row) > 1 else None
        if fecha_val is None:
            continue
        ts = _extraer_datetime(fecha_val, tiempo_val)
        if ts is not None:
            timestamps.append(ts)
    wb.close()
    if not timestamps:
        return None, None
    return min(timestamps), max(timestamps)


def cargar_datos_primarios(
    filepath: str,
    sensores: List[Sensor],
    inicio_24h: datetime,
) -> List[Sensor]:
    """Loads, merges and filters 24h data for every sensor."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    needed = set()
    for s in sensores:
        if s.col_idx_temp:
            needed.add(s.col_idx_temp)
        if s.col_idx_hum:
            needed.add(s.col_idx_hum)

    records = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        fecha_val  = row[0]
        tiempo_val = row[1]
        if fecha_val is None:
            continue
        ts = _extraer_datetime(fecha_val, tiempo_val)
        if ts is None:
            continue
        record: dict = {'timestamp': ts}
        for col_idx in needed:
            record[f'c{col_idx}'] = row[col_idx - 1] if col_idx - 1 < len(row) else None
        records.append(record)

    wb.close()

    if not records:
        return sensores

    df = pd.DataFrame(records)
    df['timestamp'] = pd.to_datetime(df['timestamp'])

    for col_idx in needed:
        col = f'c{col_idx}'
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # Merge :00/:02 duplicate rows — group by 5-min floor, first non-null per column
    df['ts_5min'] = df['timestamp'].dt.floor('5min')
    df_merged = (
        df.groupby('ts_5min', sort=True)
          .first()
          .reset_index()
          .drop(columns=['timestamp'])
          .rename(columns={'ts_5min': 'timestamp'})
    )

    fin_24h = inicio_24h + timedelta(hours=24)
    mask = (df_merged['timestamp'] >= inicio_24h) & (df_merged['timestamp'] <= fin_24h)
    df_24h = df_merged[mask].reset_index(drop=True)

    if len(df_24h) > 289:
        df_24h = df_24h.iloc[:289]

    for sensor in sensores:
        sdf = pd.DataFrame({'timestamp': df_24h['timestamp']})
        col_t = f'c{sensor.col_idx_temp}' if sensor.col_idx_temp else None
        col_h = f'c{sensor.col_idx_hum}'  if sensor.col_idx_hum  else None
        sdf['temperatura'] = df_24h[col_t].values if col_t and col_t in df_24h.columns else np.nan
        sdf['humedad']     = df_24h[col_h].values if col_h and col_h in df_24h.columns else np.nan
        sensor.datos = sdf

    return sensores
